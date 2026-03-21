"""
One-time script: Load Ticketure Excel export into BigQuery staging table.
Reads all sheets, flattens into reference.legacy_ticketure.
Only loads rows where occurred_on < 2025-07-09 (Vivenu cutover date).
"""

from __future__ import annotations

import hashlib
from datetime import datetime, date
from typing import Optional
from google.cloud import bigquery
import openpyxl

PROJECT_ID = "mercer-labs-488707"
DATASET = "reference"
TABLE = "legacy_ticketure"
XLSX_PATH = "config/Ticketure Data Export.xlsx"
VIVENU_CUTOVER = date(2025, 7, 9)

SCHEMA = [
    bigquery.SchemaField("seller", "STRING"),
    bigquery.SchemaField("event_name", "STRING"),
    bigquery.SchemaField("ticket_group", "STRING"),
    bigquery.SchemaField("ticket_type", "STRING"),
    bigquery.SchemaField("gl_code", "STRING"),
    bigquery.SchemaField("session_time", "TIMESTAMP"),
    bigquery.SchemaField("identity_name", "STRING"),
    bigquery.SchemaField("identity_email_hash", "STRING"),
    bigquery.SchemaField("cart_number", "STRING"),
    bigquery.SchemaField("order_number", "STRING"),
    bigquery.SchemaField("scan_code", "STRING"),
    bigquery.SchemaField("occurred_on", "TIMESTAMP"),
    bigquery.SchemaField("audit_action", "STRING"),
    bigquery.SchemaField("by_identity", "STRING"),
    bigquery.SchemaField("payment_gateway", "STRING"),
    bigquery.SchemaField("before_discounts_price", "FLOAT64"),
    bigquery.SchemaField("discount_amount", "FLOAT64"),
    bigquery.SchemaField("fee_fixed_inside", "FLOAT64"),
    bigquery.SchemaField("fee_percent_inside", "FLOAT64"),
    bigquery.SchemaField("fee_fixed_outside", "FLOAT64"),
    bigquery.SchemaField("fee_percent_outside", "FLOAT64"),
    bigquery.SchemaField("before_fees_price", "FLOAT64"),
    bigquery.SchemaField("after_fees_price", "FLOAT64"),
    bigquery.SchemaField("revenue", "FLOAT64"),
    bigquery.SchemaField("fee_refund_fixed", "FLOAT64"),
    bigquery.SchemaField("fee_refund_percent", "FLOAT64"),
    bigquery.SchemaField("checkout_rules", "STRING"),
    bigquery.SchemaField("code_groups", "STRING"),
    bigquery.SchemaField("codes", "STRING"),
    bigquery.SchemaField("source_sheet", "STRING"),
]

# Column indices in the Excel file
COL_MAP = {
    "seller": 0,
    "event_name": 1,
    "ticket_group": 2,
    "ticket_type": 3,
    "gl_code": 4,
    "session_time": 5,
    "identity_email": 7,
    "cart_number": 8,
    "order_number": 9,
    "scan_code": 10,
    "occurred_on": 11,
    "audit_action": 12,
    "by_identity": 13,
    "payment_gateway": 14,
    "before_discounts_price": 15,
    "discount_amount": 16,
    "fee_fixed_inside": 17,
    "fee_percent_inside": 18,
    "fee_fixed_outside": 19,
    "fee_percent_outside": 20,
    "before_fees_price": 21,
    "after_fees_price": 22,
    "revenue": 23,
    "fee_refund_fixed": 24,
    "fee_refund_percent": 25,
    "checkout_rules": 26,
    "code_groups": 27,
    "codes": 28,
}


def hash_email(email: str | None) -> str | None:
    if not email or not isinstance(email, str):
        return None
    return hashlib.sha256(email.strip().lower().encode()).hexdigest()


def to_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def to_str(val) -> str | None:
    if val is None:
        return None
    return str(val).strip() or None


def to_timestamp(val) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ["%m/%d/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y"]:
            try:
                return datetime.strptime(val.strip(), fmt)
            except ValueError:
                continue
    return None


def process_row(row: tuple, sheet_name: str) -> dict | None:
    occurred = to_timestamp(row[COL_MAP["occurred_on"]])
    if occurred and occurred.date() >= VIVENU_CUTOVER:
        return None

    return {
        "seller": to_str(row[COL_MAP["seller"]]),
        "event_name": to_str(row[COL_MAP["event_name"]]),
        "ticket_group": to_str(row[COL_MAP["ticket_group"]]),
        "ticket_type": to_str(row[COL_MAP["ticket_type"]]),
        "gl_code": to_str(row[COL_MAP["gl_code"]]),
        "session_time": to_timestamp(row[COL_MAP["session_time"]]),
        "identity_email_hash": hash_email(row[COL_MAP["identity_email"]]),
        "cart_number": to_str(row[COL_MAP["cart_number"]]),
        "order_number": to_str(row[COL_MAP["order_number"]]),
        "scan_code": to_str(row[COL_MAP["scan_code"]]),
        "occurred_on": occurred,
        "audit_action": to_str(row[COL_MAP["audit_action"]]),
        "by_identity": to_str(row[COL_MAP["by_identity"]]),
        "payment_gateway": to_str(row[COL_MAP["payment_gateway"]]),
        "before_discounts_price": to_float(row[COL_MAP["before_discounts_price"]]),
        "discount_amount": to_float(row[COL_MAP["discount_amount"]]),
        "fee_fixed_inside": to_float(row[COL_MAP["fee_fixed_inside"]]),
        "fee_percent_inside": to_float(row[COL_MAP["fee_percent_inside"]]),
        "fee_fixed_outside": to_float(row[COL_MAP["fee_fixed_outside"]]),
        "fee_percent_outside": to_float(row[COL_MAP["fee_percent_outside"]]),
        "before_fees_price": to_float(row[COL_MAP["before_fees_price"]]),
        "after_fees_price": to_float(row[COL_MAP["after_fees_price"]]),
        "revenue": to_float(row[COL_MAP["revenue"]]),
        "fee_refund_fixed": to_float(row[COL_MAP["fee_refund_fixed"]]),
        "fee_refund_percent": to_float(row[COL_MAP["fee_refund_percent"]]),
        "checkout_rules": to_str(row[COL_MAP["checkout_rules"]]),
        "code_groups": to_str(row[COL_MAP["code_groups"]]),
        "codes": to_str(row[COL_MAP["codes"]]),
        "source_sheet": sheet_name,
    }


def main():
    client = bigquery.Client(project=PROJECT_ID)
    table_ref = f"{PROJECT_ID}.{DATASET}.{TABLE}"

    # Create or replace the staging table
    table = bigquery.Table(table_ref, schema=SCHEMA)
    client.delete_table(table_ref, not_found_ok=True)
    client.create_table(table)
    print(f"Created {table_ref}")

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    total_loaded = 0
    total_skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        batch = []
        sheet_loaded = 0
        sheet_skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            record = process_row(row, sheet_name)
            if record is None:
                sheet_skipped += 1
                continue
            batch.append(record)
            sheet_loaded += 1

            # Stream in batches of 5000
            if len(batch) >= 5000:
                errors = client.insert_rows_json(table_ref, batch)
                if errors:
                    print(f"  Errors in {sheet_name}: {errors[:3]}")
                batch = []

        # Flush remaining
        if batch:
            errors = client.insert_rows_json(table_ref, batch)
            if errors:
                print(f"  Errors in {sheet_name}: {errors[:3]}")

        total_loaded += sheet_loaded
        total_skipped += sheet_skipped
        print(f"  {sheet_name}: {sheet_loaded:,} loaded, {sheet_skipped:,} skipped (post-cutover)")

    wb.close()
    print(f"\nDone. Total: {total_loaded:,} loaded, {total_skipped:,} skipped")


if __name__ == "__main__":
    main()
